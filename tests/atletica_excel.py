from pathlib import Path
import json
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- STILI EXCEL ----------

HDR_FILL  = PatternFill('solid', fgColor='1F4E79')
HDR_FONT  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
ALT_FILL  = PatternFill('solid', fgColor='EBF3FB')
WHITE_FILL = PatternFill('solid', fgColor='FFFFFF')
NORM_FONT = Font(name='Calibri', size=10)
CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT      = Alignment(horizontal='left',   vertical='center', wrap_text=True)
thin      = Side(style='thin', color='B8CCE4')
BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_cell(cell, font=None, fill=None, align=None, border=None):
    if font:  cell.font = font
    if fill:  cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = border

# ---------- ESPORTAZIONE EXCEL ----------

def make_excel(path: str, data: dict) -> None:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = 'Dettaglio Completo'
    headers = [
        'Evento ID','Sport','Categoria','Data','Destinazione',
        'Orario Ritrovo','Fine Manifestazione',
        'FIN #','Ditta','Tel Ditta','Km','Totale PAX Bus','Note Bus',
        'Istituto','Orario Partenza','Luogo Ritrovo','Persone','Rientro Presunto'
    ]
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=col, value=h)
        style_cell(c, font=HDR_FONT, fill=HDR_FILL, align=CENTER, border=BORDER)

    row_idx = 2
    alt = False
    for ev in data["events"]:
        for bg in ev["bus_groups"]:
            for i, f in enumerate(bg["fermate"]):
                fill = ALT_FILL if alt else WHITE_FILL
                values = [
                    ev["event_id"] if i == 0 else "",
                    ev["sport"] if i == 0 else "",
                    ev["categoria"] if i == 0 else "",
                    ev["data"] if i == 0 else "",
                    ev["destinazione"] if i == 0 else "",
                    ev.get("orario_ritrovo") if i == 0 else "",
                    ev.get("orario_fine_manifestazione") if i == 0 else "",
                    bg["fin"] if i == 0 else "",
                    bg.get("ditta") if i == 0 else "",
                    bg.get("ditta_tel") if i == 0 else "",
                    bg.get("km") if i == 0 else "",
                    bg.get("totale_pax") if i == 0 else "",
                    bg.get("note") if i == 0 else "",
                    f.get("istituto"),
                    f.get("orario_partenza"),
                    f.get("luogo_ritrovo"),
                    f.get("persone"),
                    f.get("rientro_presunto"),
                ]
                for col, val in enumerate(values, 1):
                    c = ws1.cell(row=row_idx, column=col, value=val)
                    align = CENTER if col in [1,2,3,4,6,7,8,11,12,15,17,18] else LEFT
                    style_cell(c, font=NORM_FONT, fill=fill, align=align, border=BORDER)
                row_idx += 1
            alt = not alt

    widths = [22,16,26,12,55,14,16,8,18,15,8,12,42,30,14,52,10,14]
    for col, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.freeze_panes = "A2"

    ws2 = wb.create_sheet('Per Istituto')
    headers2 = [
        'Sport','Data','Categoria','Destinazione',
        'FIN #','Istituto','Orario Partenza','Luogo Ritrovo','Persone','Rientro Presunto'
    ]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        style_cell(c, font=HDR_FONT, fill=HDR_FILL, align=CENTER, border=BORDER)

    rows = []
    for ev in data["events"]:
        for bg in ev["bus_groups"]:
            for f in bg["fermate"]:
                rows.append([
                    ev["sport"], ev["data"], ev["categoria"], ev["destinazione"],
                    bg["fin"], f.get("istituto"), f.get("orario_partenza"),
                    f.get("luogo_ritrovo"), f.get("persone"), f.get("rientro_presunto"),
                ])
    rows.sort(key=lambda x: (str(x[5]), str(x[1]), str(x[4])))

    for i, row in enumerate(rows, start=2):
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        for col, val in enumerate(row, 1):
            c = ws2.cell(row=i, column=col, value=val)
            align = CENTER if col in [1,2,5,7,9,10] else LEFT
            style_cell(c, font=NORM_FONT, fill=fill, align=align, border=BORDER)

    for ws in wb.worksheets:
        r = ws.max_row + 2
        ws.cell(r, 1, f"Generato il: {datetime.now():%Y-%m-%d}")

    wb.save(path)


def extract_schools_from_structured(xlsx_path: Path) -> pd.DataFrame:
    """Mirror of prepare_realSuite.extract_schools_from_structured."""
    df = pd.read_excel(xlsx_path, sheet_name="Dettaglio Completo")
    df.columns = [c.strip() for c in df.columns]
    df["FIN #"]         = df["FIN #"].ffill()
    df["Luogo Ritrovo"] = df.groupby("FIN #")["Luogo Ritrovo"].transform("ffill")
    out = pd.DataFrame({
        "Nome":         df["Istituto"].astype(str).str.strip(),
        "Indirizzo":    df["Luogo Ritrovo"].astype(str).str.strip(),
        "Partecipanti": pd.to_numeric(df["Persone"], errors="coerce").fillna(0),
    })
    out = out[out["Nome"].notna() & (out["Nome"] != "") & (out["Nome"].str.lower() != "nan")]
    out = out[out["Indirizzo"].notna() & (out["Indirizzo"] != "") & (out["Indirizzo"].str.lower() != "nan")]
    out = out.groupby(["Nome", "Indirizzo"], as_index=False).agg({"Partecipanti": "sum"})
    out["Partecipanti"] = out["Partecipanti"].astype(int)
    return out.reset_index(drop=True)


# ---------- DATI GT (Piano viaggi umano, documento .doc) ----------
# Fonte: "Atletica leggera _Cadett@_Piano viaggi_28 maggio 26.doc"
# Evento: ATLETICA LEGGERA – CADETTE/CADETTI – 28 apr 2026 – Stadio Quercia, Rovereto

data_gt = {
    "events": [
        {
            "event_id": "ATLETICA_CADETTI_280426_GT",
            "sport": "ATLETICA LEGGERA",
            "categoria": "Cadette/Cadetti",
            "data": "2026-04-28",
            "destinazione": "Stadio Quercia, Rovereto",
            "orario_ritrovo": "09:00",
            "orario_fine_manifestazione": "17:00",
            "autonomi": [],
            "bus_groups": [
                {
                    "fin": "1", "ditta": None, "ditta_tel": None, "km": 130, "totale_pax": 26, "note": None,
                    "fermate": [
                        {"istituto": "IC BORGO VALSUGANA", "orario_partenza": "07:40",
                         "luogo_ritrovo": "SSPG - Via A. Spagolla 1, Borgo Valsugana", "persone": 15, "rientro_presunto": "18:10"},
                        {"istituto": "IC LEVICO TERME",    "orario_partenza": "08:00",
                         "luogo_ritrovo": "SSPG - Via della Pace 5, Levico Terme",     "persone": 11, "rientro_presunto": "17:50"},
                    ],
                },
                {
                    "fin": "2", "ditta": None, "ditta_tel": None, "km": 188, "totale_pax": 54,
                    "note": "IC Vezzano raccolta in 2 fermate; pax stimati Cavedine 20 + Via Roma 6 = 26.",
                    "fermate": [
                        {"istituto": "IC TIONE",   "orario_partenza": "07:10",
                         "luogo_ritrovo": "SSPG – Via Circonvallazione 44, Tione", "persone": 28, "rientro_presunto": "18:45"},
                        {"istituto": "IC VEZZANO", "orario_partenza": "07:55",
                         "luogo_ritrovo": "Via XXV Aprile 7, Cavedine",            "persone": 20, "rientro_presunto": "18:00"},
                        {"istituto": "IC VEZZANO", "orario_partenza": "08:15",
                         "luogo_ritrovo": "Via Roma, Vezzano",                     "persone":  6, "rientro_presunto": "17:40"},
                    ],
                },
                {
                    "fin": "3", "ditta": None, "ditta_tel": None, "km": 72, "totale_pax": 37, "note": None,
                    "fermate": [
                        {"istituto": "IC DRO",  "orario_partenza": "07:45",
                         "luogo_ritrovo": "SSPG – Via S. Antonio 174, Dro", "persone": 11, "rientro_presunto": "18:15"},
                        {"istituto": "IC ARCO", "orario_partenza": "08:00",
                         "luogo_ritrovo": "SSPG – Via Caproni, Loc. Prabi, Arco", "persone": 26, "rientro_presunto": "18:00"},
                    ],
                },
                {
                    "fin": "4", "ditta": None, "ditta_tel": None, "km": 60, "totale_pax": 53, "note": None,
                    "fermate": [
                        {"istituto": "IC VIGOLO VATTARO",    "orario_partenza": "08:10",
                         "luogo_ritrovo": "Via Garibaldi - fermata bus, Vigolo Vattaro", "persone": 27, "rientro_presunto": "17:50"},
                        {"istituto": "IC ALDENO-MATTARELLO", "orario_partenza": "08:25",
                         "luogo_ritrovo": "Scuola – Via Torre Franca 1, Mattarello",     "persone": 13, "rientro_presunto": "17:35"},
                        {"istituto": "IC ALDENO-MATTARELLO", "orario_partenza": "08:35",
                         "luogo_ritrovo": "Fermata bus – Via Filzi, Aldeno",             "persone": 13, "rientro_presunto": "17:25"},
                    ],
                },
                {
                    "fin": "5", "ditta": None, "ditta_tel": None, "km": 92, "totale_pax": 52, "note": None,
                    "fermate": [
                        {"istituto": "IC PERGINE 1", "orario_partenza": "08:05",
                         "luogo_ritrovo": "SSPG – Via Caduti 39, Pergine Valsugana",  "persone": 26, "rientro_presunto": "17:55"},
                        {"istituto": "IC PERGINE 2", "orario_partenza": "08:10",
                         "luogo_ritrovo": "SSPG – Via Dante 276, Pergine Valsugana", "persone": 26, "rientro_presunto": "17:45"},
                    ],
                },
                {
                    "fin": "6", "ditta": None, "ditta_tel": None, "km": 51, "totale_pax": 42,
                    "note": "IC Trento 3 senza orario/luogo espliciti; eredita fermata IC Trento 6.",
                    "fermate": [
                        {"istituto": "IC TRENTO 6", "orario_partenza": "08:15",
                         "luogo_ritrovo": "Park ex Italcementi – fermata bus, Trento", "persone": 26, "rientro_presunto": "17:40"},
                        {"istituto": "IC TRENTO 3", "orario_partenza": None,
                         "luogo_ritrovo": None, "persone": 16, "rientro_presunto": None},
                    ],
                },
                {
                    "fin": "7", "ditta": None, "ditta_tel": None, "km": 54, "totale_pax": 41, "note": None,
                    "fermate": [
                        {"istituto": "IC TRENTO 2", "orario_partenza": "08:10",
                         "luogo_ritrovo": "SSPG – Via Ponte Alto 2/1, Cognola/Trento", "persone": 26, "rientro_presunto": "17:40"},
                        {"istituto": "IC TRENTO 1", "orario_partenza": "08:20",
                         "luogo_ritrovo": "SSPG – Via Znojmo 24, Povo/Trento",         "persone": 15, "rientro_presunto": "17:30"},
                    ],
                },
                {
                    "fin": "8", "ditta": None, "ditta_tel": None, "km": 46, "totale_pax": 52, "note": None,
                    "fermate": [
                        {"istituto": "IC AVIO", "orario_partenza": "08:20",
                         "luogo_ritrovo": "Viale Degasperi 69, Avio", "persone": 26, "rientro_presunto": "17:40"},
                        {"istituto": "IC ALA",  "orario_partenza": "08:35",
                         "luogo_ritrovo": "Scuola primaria – Via Betta, Ala", "persone": 26, "rientro_presunto": "17:25"},
                    ],
                },
                {
                    "fin": "9", "ditta": None, "ditta_tel": None, "km": 233, "totale_pax": 46,
                    "note": "IC Cembra in 3 fermate; pax stimati Segonzano 10 + Cembra 5 + Giovo 3 = 18.",
                    "fermate": [
                        {"istituto": "IC LADINO DI FASSA", "orario_partenza": "06:10",
                         "luogo_ritrovo": "Str. Dolomites – San Giovanni di Fassa", "persone": 12, "rientro_presunto": "19:45"},
                        {"istituto": "IC CAVALESE",        "orario_partenza": "06:50",
                         "luogo_ritrovo": "Piazza Verdi 6, Cavalese",               "persone": 16, "rientro_presunto": "19:05"},
                        {"istituto": "IC CEMBRA",          "orario_partenza": "07:30",
                         "luogo_ritrovo": "SSPG – Loc. Scancio 69, Segonzano",      "persone": 10, "rientro_presunto": "18:25"},
                        {"istituto": "IC CEMBRA",          "orario_partenza": "07:50",
                         "luogo_ritrovo": "SSPG – Via Negritelle 1, Cembra",        "persone":  5, "rientro_presunto": "18:15"},
                        {"istituto": "IC CEMBRA",          "orario_partenza": "08:10",
                         "luogo_ritrovo": "SSPG – Via Grec 2, Giovo",               "persone":  3, "rientro_presunto": "17:50"},
                    ],
                },
                {
                    "fin": "10", "ditta": None, "ditta_tel": None, "km": 70, "totale_pax": 53, "note": None,
                    "fermate": [
                        {"istituto": "IC LAVIS",    "orario_partenza": "08:10",
                         "luogo_ritrovo": "SSPG – Via Sette 13/a, Lavis",                   "persone": 26, "rientro_presunto": "17:50"},
                        {"istituto": "IC TRENTO 7", "orario_partenza": "08:25",
                         "luogo_ritrovo": "SSPG – Via Quattro Novembre 35/1, Gardolo/Trento", "persone": 27, "rientro_presunto": "17:35"},
                    ],
                },
                {
                    "fin": "11", "ditta": None, "ditta_tel": None, "km": 55, "totale_pax": 51,
                    "note": "Numerato '10' nel PDF originale (duplicato); rinumerato 11. "
                            "IC Arcivescovile senza orario/luogo espliciti; pax 19 back-calc (51-16-16).",
                    "fermate": [
                        {"istituto": "IC TRENTO 5",           "orario_partenza": "08:15",
                         "luogo_ritrovo": "Park ex Italcementi – fermata bus, Trento", "persone": 16, "rientro_presunto": "17:40"},
                        {"istituto": "IC TRENTO ARCIVESCOVILE", "orario_partenza": None,
                         "luogo_ritrovo": None, "persone": 19, "rientro_presunto": None},
                        {"istituto": "IC TRENTO 4",           "orario_partenza": "08:25",
                         "luogo_ritrovo": "Park Chiesa Sacro Cuore – Viale Verona, Trento", "persone": 16, "rientro_presunto": "17:30"},
                    ],
                },
                {
                    "fin": "12", "ditta": None, "ditta_tel": None, "km": 88, "totale_pax": 52, "note": None,
                    "fermate": [
                        {"istituto": "IC VALLE DI LEDRO", "orario_partenza": "07:45",
                         "luogo_ritrovo": "Via Falcone e Borsellino, Bezzecca",              "persone": 26, "rientro_presunto": "18:15"},
                        {"istituto": "IC RIVA 2",         "orario_partenza": "08:15",
                         "luogo_ritrovo": "Centro sportivo Malossini – Via Ginestre, Riva del Garda", "persone": 26, "rientro_presunto": "17:40"},
                    ],
                },
                {
                    "fin": "13", "ditta": None, "ditta_tel": None, "km": 127, "totale_pax": 52,
                    "note": "IC Mezzolombardo-Paganella in 3 fermate; pax stimati Andalo 26 + Spormaggiore 9 + Mezzolombardo 6 = 41.",
                    "fermate": [
                        {"istituto": "IC MEZZOLOMBARDO-PAGANELLA", "orario_partenza": "07:20",
                         "luogo_ritrovo": "Piazza San Vito 1, Andalo",         "persone": 26, "rientro_presunto": "18:20"},
                        {"istituto": "IC MEZZOLOMBARDO-PAGANELLA", "orario_partenza": "07:40",
                         "luogo_ritrovo": "Via Trento, Spormaggiore",          "persone":  9, "rientro_presunto": "18:10"},
                        {"istituto": "IC MEZZOLOMBARDO-PAGANELLA", "orario_partenza": "08:00",
                         "luogo_ritrovo": "Via degli Alpini 17, Mezzolombardo","persone":  6, "rientro_presunto": "18:00"},
                        {"istituto": "IC MEZZOCORONA",             "orario_partenza": "08:10",
                         "luogo_ritrovo": "Via Fornai 1, Mezzocorona",         "persone": 11, "rientro_presunto": "17:50"},
                    ],
                },
                {
                    "fin": "14", "ditta": None, "ditta_tel": None, "km": 65, "totale_pax": 11,
                    "note": "IC Folgaria in 2 fermate; pax stimati Lavarone 8 + Folgaria 3 = 11.",
                    "fermate": [
                        {"istituto": "IC FOLGARIA-LAVARONE-LUSERNA", "orario_partenza": "07:50",
                         "luogo_ritrovo": "Frazione Gionghi 107/6, Lavarone",     "persone": 8, "rientro_presunto": "18:00"},
                        {"istituto": "IC FOLGARIA-LAVARONE-LUSERNA", "orario_partenza": "08:20",
                         "luogo_ritrovo": "Fermata bus Palaghiaccio, Folgaria",   "persone": 3, "rientro_presunto": "17:30"},
                    ],
                },
            ],
        }
    ]
}


# ---------- DATI BUSPLAN V2 (output ottimizzatore) ----------
# Fonte: "Bus Plan_Piano_Viaggi_Atletica Leggera.docx" — prodotto da V2

data_busplan = {
    "events": [
        {
            "event_id": "ATLETICA_CADETTI_280426_BUSPLAN",
            "sport": "ATLETICA LEGGERA",
            "categoria": "Cadette/Cadetti (Bus Plan V2)",
            "data": "2026-04-28",
            "destinazione": "Stadio Quercia, Via Palestrina, Rovereto",
            "orario_ritrovo": "09:00",
            "orario_fine_manifestazione": "17:00",
            "autonomi": [],
            "bus_groups": [
                {
                    "fin": "1", "ditta": None, "ditta_tel": None, "km": 51.1, "totale_pax": 52, "note": None,
                    "fermate": [
                        {"istituto": "IC ALA",  "orario_partenza": "08:24",
                         "luogo_ritrovo": 'Scuola primaria "Abramo Betta", Via Abramo Betta 9, Ala', "persone": 26, "rientro_presunto": "17:33"},
                        {"istituto": "IC AVIO", "orario_partenza": "08:49",
                         "luogo_ritrovo": "Via Alcide Degasperi 69, 38063 Avio", "persone": 26, "rientro_presunto": "17:08"},
                    ],
                },
                {
                    "fin": "2", "ditta": None, "ditta_tel": None, "km": 49.7, "totale_pax": 49, "note": None,
                    "fermate": [
                        {"istituto": "IC TRENTO 5",             "orario_partenza": "08:28",
                         "luogo_ritrovo": "Lungadige Monte Grappa 1, 38122 Trento", "persone": 16, "rientro_presunto": "17:29"},
                        {"istituto": "IC TRENTO ARCIVESCOVILE", "orario_partenza": "08:51",
                         "luogo_ritrovo": "Trento Via Giusti cimitero, Via Giuseppe Giusti, Trento", "persone": 17, "rientro_presunto": "17:06"},
                        {"istituto": "IC TRENTO 3",             "orario_partenza": "08:56",
                         "luogo_ritrovo": "Via Vittorio Veneto, Trento", "persone": 16, "rientro_presunto": "17:01"},
                    ],
                },
                {
                    "fin": "3", "ditta": None, "ditta_tel": None, "km": 88.4, "totale_pax": 38,
                    "note": "IC Vezzano in 2 sub-gruppi alla stessa fermata (Cavedine).",
                    "fermate": [
                        {"istituto": "IC VEZZANO", "orario_partenza": "08:38",
                         "luogo_ritrovo": "Piazza Italia, Brusino, Cavedine", "persone": 12, "rientro_presunto": "18:07"},
                        {"istituto": "IC VEZZANO", "orario_partenza": "08:38",
                         "luogo_ritrovo": "Piazza Italia, Brusino, Cavedine", "persone": 12, "rientro_presunto": "17:19"},
                        {"istituto": "IC DRO",     "orario_partenza": "08:43",
                         "luogo_ritrovo": "Via Sant'Antonio 17, 38074 Dro",   "persone": 14, "rientro_presunto": "17:14"},
                    ],
                },
                {
                    "fin": "4", "ditta": None, "ditta_tel": None, "km": 156.1, "totale_pax": 15, "note": None,
                    "fermate": [
                        {"istituto": "IC BORGO VALSUGANA", "orario_partenza": "07:49",
                         "luogo_ritrovo": "Viale Trento 30/a, 38051 Grigno", "persone": 15, "rientro_presunto": "18:08"},
                    ],
                },
                {
                    "fin": "5", "ditta": None, "ditta_tel": None, "km": 85.8, "totale_pax": 44,
                    "note": "IC Cembra in 3 sub-gruppi alla stessa fermata (Giovo).",
                    "fermate": [
                        {"istituto": "IC CEMBRA", "orario_partenza": "08:43",
                         "luogo_ritrovo": "Via al Grec 2, Verla, Giovo", "persone": 6, "rientro_presunto": "17:53"},
                        {"istituto": "IC CEMBRA", "orario_partenza": "08:43",
                         "luogo_ritrovo": "Via al Grec 2, Verla, Giovo", "persone": 6, "rientro_presunto": "17:17"},
                        {"istituto": "IC CEMBRA", "orario_partenza": "08:43",
                         "luogo_ritrovo": "Via al Grec 2, Verla, Giovo", "persone": 6, "rientro_presunto": "17:14"},
                        {"istituto": "IC LAVIS",  "orario_partenza": "08:46",
                         "luogo_ritrovo": "Piazza Santi Filippo e Giacomo, Zambana, Terre d'Adige", "persone": 26, "rientro_presunto": "17:11"},
                    ],
                },
                {
                    "fin": "6", "ditta": None, "ditta_tel": None, "km": 84.8, "totale_pax": 50, "note": None,
                    "fermate": [
                        {"istituto": "IC PERGINE 2", "orario_partenza": "08:13",
                         "luogo_ritrovo": "Via delle Nazioni Unite, Maso Puller, San Cristoforo al Lago, Pergine Valsugana", "persone": 24, "rientro_presunto": "17:44"},
                        {"istituto": "IC PERGINE 1", "orario_partenza": "08:52",
                         "luogo_ritrovo": "Via Monte Cristallo, Costa, Serso, Pergine Valsugana", "persone": 26, "rientro_presunto": "17:05"},
                    ],
                },
                {
                    "fin": "7", "ditta": None, "ditta_tel": None, "km": 104.7, "totale_pax": 37,
                    "note": "IC Mezzolombardo-Paganella in 3 sub-gruppi alla stessa fermata (Spormaggiore).",
                    "fermate": [
                        {"istituto": "IC MEZZOLOMBARDO-PAGANELLA", "orario_partenza": "08:39",
                         "luogo_ritrovo": "Via Alt Spaur 5, Meano di Sopra, Spormaggiore", "persone": 10, "rientro_presunto": "18:11"},
                        {"istituto": "IC MEZZOLOMBARDO-PAGANELLA", "orario_partenza": "08:39",
                         "luogo_ritrovo": "Via Alt Spaur 5, Meano di Sopra, Spormaggiore", "persone":  6, "rientro_presunto": "17:21"},
                        {"istituto": "IC MEZZOLOMBARDO-PAGANELLA", "orario_partenza": "08:39",
                         "luogo_ritrovo": "Via Alt Spaur 5, Meano di Sopra, Spormaggiore", "persone": 10, "rientro_presunto": "17:18"},
                        {"istituto": "IC MEZZOCORONA",             "orario_partenza": "08:43",
                         "luogo_ritrovo": "Via Fornai 1, 38016 Mezzocorona",                "persone": 11, "rientro_presunto": "17:14"},
                    ],
                },
                {
                    "fin": "8", "ditta": None, "ditta_tel": None, "km": 94.9, "totale_pax": 53, "note": None,
                    "fermate": [
                        {"istituto": "IC VALLE DI LEDRO", "orario_partenza": "07:40",
                         "luogo_ritrovo": "Via Falcone e Borsellino 2, 38067 Ledro",           "persone": 26, "rientro_presunto": "18:17"},
                        {"istituto": "IC RIVA 2",         "orario_partenza": "08:32",
                         "luogo_ritrovo": 'Scuola primaria "O. Lucchi" Tenno, Via dei Laghi, Tenno', "persone": 27, "rientro_presunto": "17:25"},
                    ],
                },
                {
                    "fin": "9", "ditta": None, "ditta_tel": None, "km": 150.8, "totale_pax": 28, "note": None,
                    "fermate": [
                        {"istituto": "IC TIONE", "orario_partenza": "07:42",
                         "luogo_ritrovo": "Via Nazionale, Fontanedo, Roncone, Sella Giudicarie", "persone": 28, "rientro_presunto": "18:16"},
                    ],
                },
                {
                    "fin": "10", "ditta": None, "ditta_tel": None, "km": 128.8, "totale_pax": 27, "note": None,
                    "fermate": [
                        {"istituto": "IC TRENTO 7", "orario_partenza": "08:03",
                         "luogo_ritrovo": "Via Quattro Novembre 35, Trento", "persone": 27, "rientro_presunto": "17:55"},
                    ],
                },
                {
                    "fin": "11", "ditta": None, "ditta_tel": None, "km": 49.9, "totale_pax": 42, "note": None,
                    "fermate": [
                        {"istituto": "IC TRENTO 6", "orario_partenza": "08:30",
                         "luogo_ritrovo": "Lungadige San Nicolò, Trento",   "persone": 26, "rientro_presunto": "17:27"},
                        {"istituto": "IC TRENTO 4", "orario_partenza": "08:52",
                         "luogo_ritrovo": "Viale Verona 143, 38123 Trento", "persone": 16, "rientro_presunto": "17:05"},
                    ],
                },
                {
                    "fin": "12", "ditta": None, "ditta_tel": None, "km": 56.5, "totale_pax": 41, "note": None,
                    "fermate": [
                        {"istituto": "IC TRENTO 2", "orario_partenza": "08:19",
                         "luogo_ritrovo": "Cognola, Argentario, Martignano, Trento", "persone": 26, "rientro_presunto": "17:39"},
                        {"istituto": "IC TRENTO 1", "orario_partenza": "08:49",
                         "luogo_ritrovo": "Via Znojmo 24, 38123 Trento",             "persone": 15, "rientro_presunto": "17:08"},
                    ],
                },
                {
                    "fin": "13", "ditta": None, "ditta_tel": None, "km": 113.3, "totale_pax": 49,
                    "note": "IC Folgaria-Lavarone-Luserna in 2 sub-gruppi alla stessa fermata (Lavarone).",
                    "fermate": [
                        {"istituto": "IC LEVICO TERME",            "orario_partenza": "07:24",
                         "luogo_ritrovo": "Via della Pace 5, Levico Terme",  "persone": 11, "rientro_presunto": "18:34"},
                        {"istituto": "IC VIGOLO VATTARO",          "orario_partenza": "08:13",
                         "luogo_ritrovo": "Vigolo Vattaro, Comunità Alta Valsugana", "persone": 27, "rientro_presunto": "17:45"},
                        {"istituto": "IC FOLGARIA-LAVARONE-LUSERNA", "orario_partenza": "08:57",
                         "luogo_ritrovo": "Scuola Primaria Lavarone, Piazza Maria Teresa d'Austria, Lavarone", "persone": 5, "rientro_presunto": "17:25"},
                        {"istituto": "IC FOLGARIA-LAVARONE-LUSERNA", "orario_partenza": "08:57",
                         "luogo_ritrovo": "Scuola Primaria Lavarone, Piazza Maria Teresa d'Austria, Lavarone", "persone": 6, "rientro_presunto": "17:00"},
                    ],
                },
                {
                    "fin": "14", "ditta": None, "ditta_tel": None, "km": 215.6, "totale_pax": 28, "note": None,
                    "fermate": [
                        {"istituto": "IC LADINO DI FASSA", "orario_partenza": "06:59",
                         "luogo_ritrovo": "Via Riccardo Lowy, Moena", "persone": 12, "rientro_presunto": "18:58"},
                        {"istituto": "IC CAVALESE",        "orario_partenza": "08:31",
                         "luogo_ritrovo": "Via Lagorai, 38099 Cavalese", "persone": 16, "rientro_presunto": "17:26"},
                    ],
                },
                {
                    "fin": "15", "ditta": None, "ditta_tel": None, "km": 71.8, "totale_pax": 52, "note": None,
                    "fermate": [
                        {"istituto": "IC ARCO",           "orario_partenza": "07:36",
                         "luogo_ritrovo": "Via Paolina Caproni Maini 28, Arco", "persone": 26, "rientro_presunto": "18:21"},
                        {"istituto": "IC ALDENO-MATTARELLO", "orario_partenza": "08:10",
                         "luogo_ritrovo": "Via Venticinque Aprile, Aldeno",    "persone": 26, "rientro_presunto": "17:47"},
                    ],
                },
            ],
        }
    ]
}


# ---------- CONFIG ----------

config = {
    "destination": "Stadio Quercia, Via Palestrina, Rovereto",
    "capacity": 54,
    "orario_fine_manifestazione": "17:00",
    "destination_lat": 45.8862,
    "destination_lon": 11.0363,
}


# ---------- MAIN ----------

REALSUITE_DIR = Path(__file__).parent / "realSuite" / "Atletica Rovereto"

if __name__ == "__main__":
    REALSUITE_DIR.mkdir(parents=True, exist_ok=True)

    # 1. groundtruth.xlsx (piano umano GT)
    gt_path = REALSUITE_DIR / "groundtruth.xlsx"
    make_excel(str(gt_path), data_gt)
    print(f"✓ groundtruth.xlsx → {gt_path}")

    # 2. busplan_output.xlsx (output V2 per confronto visivo)
    bp_path = REALSUITE_DIR / "busplan_output.xlsx"
    make_excel(str(bp_path), data_busplan)
    print(f"✓ busplan_output.xlsx → {bp_path}")

    # 3. input.xlsx estratto dal GT
    schools_df = extract_schools_from_structured(gt_path)
    inp_path = REALSUITE_DIR / "input.xlsx"
    schools_df.to_excel(str(inp_path), index=False)
    print(f"✓ input.xlsx → {inp_path}  ({len(schools_df)} fermate)")

    # 4. config.json
    cfg_path = REALSUITE_DIR / "config.json"
    cfg_path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"✓ config.json → {cfg_path}")

    print("\nProssimi step:")
    print("  python tests/prepare_realSuite.py --correct   # corregge indirizzi via AI")
    print("  python tests/prepare_realSuite.py --geocode   # geocodifica + distance_matrix")
