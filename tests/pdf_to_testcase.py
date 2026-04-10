from pathlib import Path
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# --- Stili base ---
HDR_FILL  = PatternFill('solid', fgColor='1F4E79')
HDR_FONT  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
ALT_FILL  = PatternFill('solid', fgColor='EBF3FB')
WHITE_FILL = PatternFill('solid', fgColor='FFFFFF')
NORM_FONT = Font(name='Calibri', size=10)
CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT      = Alignment(horizontal='left',   vertical='center', wrap_text=True)
thin      = Side(style='thin', color='B8CCE4')
BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_cell(cell, font=None, fill=None, align=None, border=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if border:
        cell.border = border


def make_excel(path: str, data: dict) -> None:
    """
    Crea l'Excel di ground truth a partire dalla struttura `data`.

    `data` deve rispettare lo schema:
      data = {"events": [ { event_id, sport, ..., bus_groups: [ {fin,..., fermate:[...]} ] } ]}
    """
    wb = Workbook()

    # ---------------- Foglio 1: Dettaglio Completo ----------------
    ws1 = wb.active
    ws1.title = 'Dettaglio Completo'

    headers = [
        'Evento ID','Sport','Categoria','Data','Destinazione',
        'Orario Ritrovo','Fine Manifestazione',
        'FIN #','Ditta','Tel Ditta','Km','Totale PAX Bus','Note Bus',
        'Istituto','Orario Partenza','Luogo Ritrovo','Persone','Rientro Presunto'
    ]
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=col, value=h)
        style_cell(c, font=HDR_FONT, fill=HDR_FILL, align=CENTER, border=BORDER)

    row_idx = 2
    alt = False

    for ev in data["events"]:
        for bg in ev["bus_groups"]:
            for i, f in enumerate(bg["fermate"]):
                fill = ALT_FILL if alt else WHITE_FILL

                values = [
                    ev["event_id"] if i == 0 else '',
                    ev["sport"]     if i == 0 else '',
                    ev["categoria"] if i == 0 else '',
                    ev["data"]      if i == 0 else '',
                    ev["destinazione"] if i == 0 else '',
                    ev.get("orario_ritrovo") if i == 0 else '',
                    ev.get("orario_fine_manifestazione") if i == 0 else '',
                    bg["fin"]       if i == 0 else '',
                    bg.get("ditta") if i == 0 else '',
                    bg.get("ditta_tel") if i == 0 else '',
                    bg.get("km")    if i == 0 else '',
                    bg.get("totale_pax") if i == 0 else '',
                    bg.get("note")  if i == 0 else '',
                    f.get("istituto"),
                    f.get("orario_partenza"),
                    f.get("luogo_ritrovo"),
                    f.get("persone"),
                    f.get("rientro_presunto"),
                ]

                for col, val in enumerate(values, 1):
                    c = ws1.cell(row=row_idx, column=col, value=val)
                    align = CENTER if col in [1,2,3,4,6,7,8,11,12,15,17,18] else LEFT
                    style_cell(c, font=NORM_FONT, fill=fill, align=align, border=BORDER)

                row_idx += 1

            alt = not alt

    widths = [22,16,26,12,55,14,16,8,18,15,8,12,42,30,14,52,10,14]
    for col, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.freeze_panes = "A2"

    # ---------------- Foglio 2: Per Istituto ----------------
    ws2 = wb.create_sheet('Per Istituto')

    headers2 = [
        'Sport','Data','Categoria','Destinazione',
        'FIN #','Istituto','Orario Partenza','Luogo Ritrovo','Persone','Rientro Presunto'
    ]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        style_cell(c, font=HDR_FONT, fill=HDR_FILL, align=CENTER, border=BORDER)

    rows = []
    for ev in data["events"]:
        for bg in ev["bus_groups"]:
            for f in bg["fermate"]:
                rows.append([
                    ev["sport"],
                    ev["data"],
                    ev["categoria"],
                    ev["destinazione"],
                    bg["fin"],
                    f.get("istituto"),
                    f.get("orario_partenza"),
                    f.get("luogo_ritrovo"),
                    f.get("persone"),
                    f.get("rientro_presunto"),
                ])

    # ordina per istituto, poi data, poi FIN
    rows.sort(key=lambda x: (str(x[5]), str(x[1]), str(x[4])))

    for i, row in enumerate(rows, start=2):
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        for col, val in enumerate(row, 1):
            c = ws2.cell(row=i, column=col, value=val)
            align = CENTER if col in [1,2,5,7,9,10] else LEFT
            style_cell(c, font=NORM_FONT, fill=fill, align=align, border=BORDER)

    # footer
    for ws in wb.worksheets:
        r = ws.max_row + 2
        ws.cell(r, 1, f"Generato il: {datetime.now():%Y-%m-%d} | Fonte: estrazione da PDF")

    wb.save(path)


def save_pair(base_name: str, data: dict, out_dir: str = "output"):
    """
    Salva JSON + Excel con lo stesso nome base.

    - base_name: nome base del file (senza estensione PDF)
    - data: struttura `data`
    - out_dir: directory di output
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    json_path = out_dir / f"{base_name}_structured.json"
    xlsx_path = out_dir / f"{base_name}_structured.xlsx"

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    make_excel(str(xlsx_path), data)

    return str(json_path), str(xlsx_path)